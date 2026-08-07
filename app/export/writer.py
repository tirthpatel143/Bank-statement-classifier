import io
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from typing import List, Dict, Any, Optional, Union
from app.schemas.models import StatementProcessResult, Transaction

def export_to_excel(result: StatementProcessResult) -> bytes:
    """
    Generates a multi-sheet formatted Excel workbook containing:
    Sheet 1: Transactions (with Sender & Recipient columns)
    Sheet 2: Account Details
    Sheet 3: Validation Report
    Sheet 4: Category Summary
    """
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    border_side = Side(border_style="thin", color="D9D9D9")
    thin_border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")
    right_align = Alignment(horizontal="right", vertical="center")

    # -------------------------------------------------------------
    # Sheet 1: Transactions
    # -------------------------------------------------------------
    ws_tx = wb.create_sheet(title="Transactions")
    tx_headers = [
        "Date", "Description", "Sender (Money From)", "Recipient (Money To)", 
        "Debit", "Credit", "Balance", "Category", "Method", "Status"
    ]
    ws_tx.append(tx_headers)

    for cell in ws_tx[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align

    for tx in result.transactions:
        method = "Rule" if tx.classification_method == "rule" else ("ML model" if tx.classification_method == "ml" else "Manual")
        status = "Valid" if (tx.row_valid and tx.balance_check and not tx.needs_review) else "Needs Review"
        row = [
            tx.date,
            tx.description,
            tx.sender or "Self",
            tx.recipient or "Self",
            tx.debit if tx.debit is not None else "",
            tx.credit if tx.credit is not None else "",
            tx.balance,
            tx.category or "Other",
            method,
            status
        ]
        ws_tx.append(row)

    for row in ws_tx.iter_rows(min_row=2, max_row=ws_tx.max_row, min_col=1, max_col=10):
        row[0].alignment = center_align
        row[1].alignment = left_align
        row[2].alignment = left_align
        row[3].alignment = left_align
        row[4].number_format = "#,##0.00"
        row[4].alignment = right_align
        row[5].number_format = "#,##0.00"
        row[5].alignment = right_align
        row[6].number_format = "#,##0.00"
        row[6].alignment = right_align
        row[7].alignment = center_align
        row[8].alignment = center_align
        row[9].alignment = center_align
        for cell in row:
            cell.border = thin_border

    # Auto-fit column widths
    for col in ws_tx.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws_tx.column_dimensions[col_letter].width = max(max_len + 3, 12)

    # -------------------------------------------------------------
    # Sheet 2: Account Details
    # -------------------------------------------------------------
    ws_acc = wb.create_sheet(title="Account Details")
    ws_acc.append(["Field", "Value"])
    for cell in ws_acc[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align

    acc = result.account_details
    acc_rows = [
        ["Bank Name", acc.bank_name.value or "Unknown"],
        ["Account Holder", acc.account_holder.value or "Unknown"],
        ["Masked Account Number", acc.masked_account_number],
        ["IFSC Code", acc.ifsc.value or "N/A"],
        ["Branch", acc.branch.value or "N/A"],
        ["Statement Period Start", acc.statement_period_start or "N/A"],
        ["Statement Period End", acc.statement_period_end or "N/A"],
        ["Opening Balance", acc.opening_balance if acc.opening_balance is not None else "N/A"],
        ["Closing Balance", acc.closing_balance if acc.closing_balance is not None else "N/A"]
    ]
    for r in acc_rows:
        ws_acc.append(r)

    for row in ws_acc.iter_rows(min_row=2, max_row=ws_acc.max_row, min_col=1, max_col=2):
        row[0].alignment = left_align
        row[1].alignment = left_align
        for cell in row:
            cell.border = thin_border
    ws_acc.column_dimensions['A'].width = 25
    ws_acc.column_dimensions['B'].width = 35

    # -------------------------------------------------------------
    # Sheet 3: Validation Report
    # -------------------------------------------------------------
    ws_val = wb.create_sheet(title="Validation Report")
    ws_val.append(["Check Name", "Status", "Details"])
    for cell in ws_val[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align

    vr = result.validation_report
    val_rows = [
        ["Total Processed Rows", "Info", vr.total_rows],
        ["Valid Rows", "Pass" if vr.valid_rows == vr.total_rows else "Review", vr.valid_rows],
        ["Rows Needing Review", "Warning" if vr.invalid_rows > 0 else "Pass", vr.invalid_rows],
        ["Balance Pass Rate", "Pass" if vr.balance_check_pass_rate >= 95 else "Warning", f"{vr.balance_check_pass_rate}%"],
        ["Duplicate Rows Flagged", "Info", vr.duplicate_count],
        ["Opening Balance Reconciled", "Pass" if vr.opening_balance_matched else "Warning", str(vr.opening_balance_matched)],
        ["Closing Balance Reconciled", "Pass" if vr.closing_balance_matched else "Warning", str(vr.closing_balance_matched)]
    ]
    for r in val_rows:
        ws_val.append(r)

    for row in ws_val.iter_rows(min_row=2, max_row=ws_val.max_row, min_col=1, max_col=3):
        row[0].alignment = left_align
        row[1].alignment = center_align
        row[2].alignment = left_align
        for cell in row:
            cell.border = thin_border
    ws_val.column_dimensions['A'].width = 30
    ws_val.column_dimensions['B'].width = 15
    ws_val.column_dimensions['C'].width = 40

    # -------------------------------------------------------------
    # Sheet 4: Category Summary
    # -------------------------------------------------------------
    ws_cat = wb.create_sheet(title="Category Summary")
    ws_cat.append(["Category", "Transaction Count", "Total Debit (₹)", "Total Credit (₹)"])
    for cell in ws_cat[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align

    for cs in result.classification_summary:
        ws_cat.append([
            cs.category,
            cs.count,
            cs.total_debit,
            cs.total_credit
        ])

    for row in ws_cat.iter_rows(min_row=2, max_row=ws_cat.max_row, min_col=1, max_col=4):
        row[0].alignment = left_align
        row[1].alignment = center_align
        row[2].number_format = "#,##0.00"
        row[2].alignment = right_align
        row[3].number_format = "#,##0.00"
        row[3].alignment = right_align
        for cell in row:
            cell.border = thin_border

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()

def export_to_csv(result_or_txs: Union[StatementProcessResult, List[Transaction]]) -> str:
    """
    Generates CSV string representation of extracted transactions with Sender & Recipient columns.
    Accepts either StatementProcessResult or List[Transaction].
    """
    txs = result_or_txs.transactions if hasattr(result_or_txs, "transactions") else result_or_txs
    data = []
    for tx in txs:
        method = "Rule" if tx.classification_method == "rule" else ("ML model" if tx.classification_method == "ml" else "Manual")
        status = "Valid" if (tx.row_valid and tx.balance_check and not tx.needs_review) else "Needs Review"
        data.append({
            "date": tx.date,
            "description": tx.description,
            "sender": tx.sender or "Self",
            "recipient": tx.recipient or "Self",
            "debit": tx.debit if tx.debit is not None else "",
            "credit": tx.credit if tx.credit is not None else "",
            "balance": tx.balance,
            "category": tx.category or "Other",
            "classification_method": method,
            "status": status
        })
    df = pd.DataFrame(data)
    return df.to_csv(index=False)
